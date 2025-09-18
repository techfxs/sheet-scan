from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import pandas as pd
import io
from openpyxl import load_workbook
import time

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/uploadwithpandas")
async def upload_file(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        # Example: Add error messages in a new column
        import re
        def summarize_row_errors(row):
            errors = []
            for col in df.columns:
                val = row[col]
                if pd.isnull(val):
                    continue
                if re.search(r'[A-Za-z]', str(val)):
                    errors.append(f"{col}: contains alphabets")
            return "; ".join(errors) if errors else ''
        df['ValidationErrors'] = df.apply(summarize_row_errors, axis=1)
        output = io.BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
            'Content-Disposition': 'attachment; filename=processed.xlsx'
        })
    except Exception as e:
        return {"error": str(e)}


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    start_time = time.time()
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Only process up to 100,000 rows (excluding header)
        max_rows = min(ws.max_row, 100_001)  # 1 for header, 100,000 for data

        # Set header for column T (20th column)
        ws.cell(row=1, column=20, value="ValidationErrors")

        import re
        for row_idx in range(2, max_rows + 1):  # Rows 2 to 100001 (if present)
            errors = []
            for col_idx in range(1, 20):  # Columns A to S (1 to 19)
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and re.search(r'[A-Za-z]', str(cell.value)):
                    col_name = ws.cell(row=1, column=col_idx).value
                    errors.append(f"{col_name}: contains alphabets")
            ws.cell(row=row_idx, column=20, value="; ".join(errors) if errors else "")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        elapsed = time.time() - start_time
        print(f"Processing time (openpyxl): {elapsed:.2f} seconds")
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=processed.xlsx'}
        )
    except Exception as e:
        return {"error": str(e)}


@app.post("/uploadcsv2")
async def upload_csv(file: UploadFile = File(...)):
    start_time = time.time()
    try:
        contents = await file.read()
        # Read CSV into pandas DataFrame
        df = pd.read_csv(io.BytesIO(contents))

        import re
        # Only process columns A to S (first 19 columns)
        columns_to_check = df.columns[:19]
        def summarize_row_errors(row):
            errors = []
            for col in columns_to_check:
                val = row[col]
                if pd.isnull(val):
                    continue
                if re.search(r'[A-Za-z]', str(val)):
                    errors.append(f"{col}: contains alphabets")
            return "; ".join(errors) if errors else ''
        # Write summary in column T (20th column)
        df['ValidationErrors'] = df.apply(summarize_row_errors, axis=1)

        output = io.BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        elapsed = time.time() - start_time
        print(f"Processing time (csv): {elapsed:.2f} seconds")
        return StreamingResponse(
            output,
            media_type='text/csv',
            headers={'Content-Disposition': 'attachment; filename=processed.csv'}
        )
    except Exception as e:
        return {"error": str(e)}



